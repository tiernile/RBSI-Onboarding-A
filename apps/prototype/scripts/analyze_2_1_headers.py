#!/usr/bin/env python3
"""
Header Analysis for 2.1 Spreadsheet using openpyxl

Analyzes the new spreadsheet to identify structure and new Nile suggestion columns.
"""
import json
from pathlib import Path
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = APP_DIR / 'data'
INCOMING = DATA_DIR / 'incoming' / '20250916-master-spreadsheet-2.1.xlsx'

def column_letter_to_number(col_letter):
    """Convert column letter to number (A=1, B=2, etc.)"""
    num = 0
    for c in col_letter:
        num = num * 26 + ord(c) - ord('A') + 1
    return num

def analyze_spreadsheet():
    """Analyze the 2.1 spreadsheet structure"""
    print(f"Analyzing: {INCOMING}")
    
    if not INCOMING.exists():
        print(f"ERROR: File not found: {INCOMING}")
        return
    
    # Load workbook
    wb = load_workbook(INCOMING, data_only=True)
    print(f"Worksheets: {wb.sheetnames}")
    
    # Focus on LP Proposal sheet
    if 'LP Proposal' in wb.sheetnames:
        ws = wb['LP Proposal']
        print(f"\nAnalyzing 'LP Proposal' sheet...")
        print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Check first few rows to find headers
        print("\n=== SCANNING FOR HEADERS ===")
        for row_num in range(1, 6):
            print(f"\nRow {row_num}:")
            row_headers = {}
            for col_num in range(1, min(50, ws.max_column + 1)):  # Check first 50 columns
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    col_letter = cell.column_letter
                    value = str(cell.value).strip()
                    row_headers[col_letter] = value
                    if len(value) < 100:  # Only show reasonable length values
                        print(f"  {col_letter}: {value}")
            
            # Check if this looks like a header row
            if row_headers:
                header_keywords = ['keyname', 'field name', 'data type', 'nile suggested', 'paul']
                keyword_count = sum(1 for keyword in header_keywords 
                                  for header in row_headers.values() 
                                  if keyword in header.lower())
                print(f"  Header keywords found: {keyword_count}")
                
                if keyword_count >= 3:  # Likely header row
                    print(f"  *** ROW {row_num} LOOKS LIKE HEADERS ***")
                    
                    # Now do detailed analysis
                    print(f"\n=== DETAILED ANALYSIS OF ROW {row_num} ===")
                    
                    # New Nile suggestion columns (L, M, N)
                    nile_cols = {}
                    for col, header in row_headers.items():
                        header_lower = header.lower()
                        if 'nile suggested field name' in header_lower:
                            nile_cols['field_name'] = col
                            print(f"✓ Nile Suggested Field Name: {col} = '{header}'")
                        elif 'nile suggested description' in header_lower:
                            nile_cols['description'] = col  
                            print(f"✓ Nile Suggested Description: {col} = '{header}'")
                        elif 'nile suggested section' in header_lower:
                            nile_cols['section'] = col
                            print(f"✓ Nile Suggested Section: {col} = '{header}'")
                    
                    # Core mapping columns
                    mapping_cols = {}
                    for col, header in row_headers.items():
                        header_lower = header.lower()
                        if header_lower == 'keyname':
                            mapping_cols['id'] = col
                            print(f"✓ ID/Keyname: {col}")
                        elif header_lower == 'field name':
                            mapping_cols['label'] = col
                            print(f"✓ Field Name: {col}")
                        elif 'visibility condition' in header_lower or 'group name' in header_lower:
                            mapping_cols['visibility'] = col
                            print(f"✓ Visibility: {col} = '{header}'")
                        elif header_lower == 'data type':
                            mapping_cols['data_type'] = col
                            print(f"✓ Data Type: {col}")
                    
                    # Additional important columns
                    other_cols = {}
                    for col, header in row_headers.items():
                        header_lower = header.lower()
                        if 'live question order' in header_lower:
                            other_cols['live_order'] = col
                            print(f"✓ Live Question Order: {col}")
                        elif 'paul question order' in header_lower:
                            other_cols['paul_order'] = col
                            print(f"✓ PAUL Question Order: {col}")
                        elif 'paul section suggestion' in header_lower:
                            other_cols['paul_section'] = col
                            print(f"✓ PAUL Section Suggestion: {col}")
                        elif header_lower == 'internal':
                            other_cols['internal'] = col
                            print(f"✓ Internal: {col}")
                        elif header_lower == 'action':
                            other_cols['action'] = col
                            print(f"✓ Action: {col}")
                    
                    # Sample data from next few rows
                    print(f"\n=== DATA SAMPLE (after header row {row_num}) ===")
                    for sample_row in range(row_num + 1, min(row_num + 4, ws.max_row + 1)):
                        sample_data = {}
                        key_cols = ['A', 'B'] + list(nile_cols.values())[:3]
                        for col in key_cols:
                            if col in row_headers:
                                cell_value = ws[f"{col}{sample_row}"].value
                                if cell_value:
                                    sample_data[f"{col}({row_headers[col][:20]})"] = str(cell_value)[:50]
                        
                        if sample_data:
                            print(f"  Row {sample_row}: {sample_data}")
                    
                    # Generate suggested mapping
                    print(f"\n=== SUGGESTED v2.1 MAPPING ===")
                    suggested_cols = {
                        "id": mapping_cols.get('id', 'KEYNAME'),
                        "label": mapping_cols.get('label', 'FIELD NAME'),
                        "visibility": mapping_cols.get('visibility', 'VISIBILITY CONDITION/GROUP NAME'),
                        "data_type": mapping_cols.get('data_type', 'DATA TYPE'),
                        **{f"nile_{k}": v for k, v in nile_cols.items()},
                        **{f"paul_{k}": v for k, v in other_cols.items() if 'paul' in k.lower()}
                    }
                    
                    for key, col in suggested_cols.items():
                        if col in row_headers:
                            print(f"  {key}: {col} = '{row_headers[col]}'")
                    
                    break  # Found headers, stop scanning
    
    wb.close()

if __name__ == "__main__":
    analyze_spreadsheet()